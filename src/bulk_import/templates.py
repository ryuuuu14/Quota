import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from database import get_connection

ALLOWED_LOAI = [
    "LT", "TH", "NN_CNTT",
    "THẠC SĨ", "TIẾN SĨ",
    "LLCT TRUNG CẤP", "LLCT CAO CẤP",
    "BỒI DƯỠNG"
]

def generate_excel_template(timeframe_name):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.id, t.name,
               (SELECT value_text FROM teacher_role_history
                WHERE teacher_id = t.id AND record_type = 'TITLE'
                ORDER BY start_date DESC LIMIT 1) as title,
               (SELECT value_text FROM teacher_role_history
                WHERE teacher_id = t.id AND record_type = 'DEPARTMENT'
                ORDER BY start_date DESC LIMIT 1) as dept
        FROM teachers t
        WHERE t.employment_type IN ('TEACHER', 'STAFF')
        ORDER BY t.name
    """)
    teachers = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhập giờ giảng"

    ws.protection.sheet = True
    ws.protection.enable()

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="1F5F3F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    instruction_font = Font(name=font_family, size=10, italic=True, color="555555")

    header_fill = PatternFill(start_color="1F5F3F", end_color="1F5F3F", fill_type="solid")
    instruction_fill = PatternFill(start_color="EAF2EC", end_color="EAF2EC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    ws.merge_cells("A1:K1")
    ws["A1"] = f"MẪU NHẬP GIỜ GIẢNG CHI TIẾT - {timeframe_name.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:K2")
    ws["A2"] = ("Hướng dẫn: Điền thông tin từng lớp/môn giảng dạy vào các ô trắng. "
                "Cột Mã GV, Họ tên, Chức danh, Đơn vị được bảo vệ. "
                "Cột Loại: LT, TH, NN_CNTT, THẠC SĨ, TIẾN SĨ, LLCT TRUNG CẤP, LLCT CAO CẤP, BỒI DƯỠNG.")
    ws["A2"].font = instruction_font
    ws["A2"].fill = instruction_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Mã GV (Khóa)", "Họ tên (Khóa)", "Chức danh (Khóa)", "Đơn vị (Khóa)",
        "Tên môn học", "Loại", "Nhóm", "Sỉ số", "Tiết quy đổi",
        "Hệ số tín chỉ", "Ghi chú"
    ]

    ws.row_dimensions[4].height = 30
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    start_row = 5
    for idx, t in enumerate(teachers):
        current_row = start_row + idx
        ws.row_dimensions[current_row].height = 22

        locked_vals = [t['id'], t['name'], t['title'] or '', t['dept'] or '']
        for col_idx, val in enumerate(locked_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.protection = Protection(locked=True)

        for col in range(5, 12):
            cell = ws.cell(row=current_row, column=col)
            if col in (8, 9, 10):
                cell.value = 0
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = ""
                cell.alignment = Alignment(vertical="center")
            cell.font = data_font
            cell.border = thin_border
            cell.protection = Protection(locked=False)

    loai_col_letter = get_column_letter(6)
    loai_range = f"{loai_col_letter}5:{loai_col_letter}{start_row + len(teachers) - 1}"
    loai_formula = ','.join(ALLOWED_LOAI)
    dv = DataValidation(type="list", formula1=f'"{loai_formula}"', allow_blank=True)
    dv.error = "Giá trị không hợp lệ. Chọn: LT, TH, NN_CNTT, THẠC SĨ, TIẾN SĨ, LLCT TRUNG CẤP, LLCT CAO CẤP, BỒI DƯỠNG"
    dv.errorTitle = "Loại không hợp lệ"
    ws.add_data_validation(dv)
    dv.add(loai_range)

    col_widths = {1: 10, 2: 22, 3: 18, 4: 22, 5: 22, 6: 16, 7: 8, 8: 8, 9: 12, 10: 12, 11: 16}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
