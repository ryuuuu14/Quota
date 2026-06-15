import os
import openpyxl
import io
from pipeline.templates import generate_teachers_template, generate_activities_template

def test_generate_teachers_template():
    data = generate_teachers_template("Chính trị, Pháp luật, Nghiệp vụ")
    assert isinstance(data, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Thông tin nhà giáo"
    # Header check
    assert ws.cell(row=4, column=1).value == "Mã GV"
    assert ws.cell(row=4, column=2).value == "Họ tên"
    assert ws.cell(row=4, column=3).value == "Đơn vị công tác"

def test_generate_activities_template():
    data = generate_activities_template("Chính trị, Pháp luật, Nghiệp vụ", "2024-2025")
    assert isinstance(data, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Hoạt động chuyên môn"
    assert ws.cell(row=4, column=1).value == "Mã GV"
    assert ws.cell(row=4, column=2).value == "Tên loại hoạt động"
