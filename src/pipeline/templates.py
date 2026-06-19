import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from database import get_connection

ALLOWED_LOAI = [
    "LT",
    "TH",
    "NN_CNTT",
    "THẠC SĨ",
    "TIẾN SĨ",
    "LLCT TRUNG CẤP",
    "LLCT CAO CẤP",
    "BỒI DƯỠNG",
]


def generate_excel_template(timeframe_name):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.id, t.name,
               (SELECT value_text FROM teacher_role_history
                WHERE teacher_id = t.id AND record_type = 'TITLE'
                ORDER BY start_date DESC LIMIT 1) as title,
               (SELECT value_text FROM teacher_role_history
                WHERE teacher_id = t.id AND record_type = 'DEPARTMENT'
                ORDER BY start_date DESC LIMIT 1) as dept
        FROM teachers t
        WHERE t.employment_type IN ('TEACHER', 'STAFF')
        ORDER BY t.name
    """)
    teachers = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhập giờ giảng"

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="1F5F3F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    instruction_font = Font(name=font_family, size=10, italic=True, color="555555")

    header_fill = PatternFill(
        start_color="1F5F3F", end_color="1F5F3F", fill_type="solid"
    )
    instruction_fill = PatternFill(
        start_color="EAF2EC", end_color="EAF2EC", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:K1")
    ws["A1"] = f"MẪU NHẬP GIỜ GIẢNG CHI TIẾT - {timeframe_name.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "Hướng dẫn: Điền thông tin từng lớp/môn giảng dạy vào các ô trắng. "
        "Cột Mã GV, Họ tên, Chức danh, Đơn vị được bảo vệ. "
        "Cột Loại: LT, TH, NN_CNTT, THẠC SĨ, TIẾN SĨ, LLCT TRUNG CẤP, LLCT CAO CẤP, BỒI DƯỠNG."
    )
    ws["A2"].font = instruction_font
    ws["A2"].fill = instruction_fill
    ws["A2"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[2].height = 30

    headers = [
        "Mã GV (Khóa)",
        "Họ tên (Khóa)",
        "Chức danh (Khóa)",
        "Đơn vị (Khóa)",
        "Tên môn học",
        "Loại",
        "Nhóm",
        "Sỉ số",
        "Tiết quy đổi",
        "Hệ số tín chỉ",
        "Ghi chú",
    ]

    ws.row_dimensions[4].height = 30
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    start_row = 5
    for idx, t in enumerate(teachers):
        current_row = start_row + idx
        ws.row_dimensions[current_row].height = 22

        locked_vals = [t["id"], t["name"], t["title"] or "", t["dept"] or ""]
        for col_idx, val in enumerate(locked_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.protection = Protection(locked=True)

        for col in range(5, 12):
            cell = ws.cell(row=current_row, column=col)
            if col in (8, 9, 10):
                cell.value = 0
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = ""
                cell.alignment = Alignment(vertical="center")
            cell.font = data_font
            cell.border = thin_border
            cell.protection = Protection(locked=False)

    loai_col_letter = get_column_letter(6)
    loai_range = f"{loai_col_letter}5:{loai_col_letter}{start_row + len(teachers) - 1}"
    loai_formula = ",".join(ALLOWED_LOAI)
    dv = DataValidation(type="list", formula1=f'"{loai_formula}"', allow_blank=True)
    dv.error = "Giá trị không hợp lệ. Chọn: LT, TH, NN_CNTT, THẠC SĨ, TIẾN SĨ, LLCT TRUNG CẤP, LLCT CAO CẤP, BỒI DƯỠNG"
    dv.errorTitle = "Loại không hợp lệ"
    ws.add_data_validation(dv)
    dv.add(loai_range)

    col_widths = {
        1: 10,
        2: 22,
        3: 18,
        4: 22,
        5: 22,
        6: 16,
        7: 8,
        8: 8,
        9: 12,
        10: 12,
        11: 16,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_teachers_template(dept_name):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT t.id as teacher_id, t.name,
               (SELECT value_text FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'TITLE' ORDER BY start_date DESC LIMIT 1) as title,
               (SELECT start_date FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'TITLE' ORDER BY start_date DESC LIMIT 1) as title_date,
               (SELECT value_text FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'ROLE' ORDER BY start_date DESC LIMIT 1) as role,
               (SELECT start_date FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'ROLE' ORDER BY start_date DESC LIMIT 1) as role_date
        FROM teachers t
        WHERE (SELECT value_text FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'DEPARTMENT' ORDER BY start_date DESC LIMIT 1) = ?
        ORDER BY t.name
    """,
        (dept_name,),
    )
    teachers = [dict(r) for r in cursor.fetchall()]

    # Fetch options for dropdown validations
    cursor.execute("SELECT name FROM titles ORDER BY name")
    titles_db = [r["name"] for r in cursor.fetchall()]

    cursor.execute(
        "SELECT name FROM reduction_rules WHERE rule_type = 'ROLE' ORDER BY name"
    )
    roles_db = [r["name"] for r in cursor.fetchall()]
    conn.close()

    if not titles_db:
        titles_db = ["Chưa cài đặt"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thông tin nhà giáo"

    # Populate metadata sheet for lists
    meta_ws = wb.create_sheet(title="Metadata")
    meta_ws.sheet_state = "hidden"

    for idx, val in enumerate(titles_db, 1):
        meta_ws.cell(row=idx, column=1, value=val)

    roles_list_db = ["Không có"] + roles_db
    for idx, val in enumerate(roles_list_db, 1):
        meta_ws.cell(row=idx, column=3, value=val)

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="1F5F3F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    instruction_font = Font(name=font_family, size=10, italic=True, color="555555")

    header_fill = PatternFill(
        start_color="1F5F3F", end_color="1F5F3F", fill_type="solid"
    )
    instruction_fill = PatternFill(
        start_color="EAF2EC", end_color="EAF2EC", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    last_col = get_column_letter(10)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"MẪU NHẬP CÁN BỘ HÀNG LOẠT - ĐƠN VỊ: {dept_name.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Hướng dẫn: Điền thông tin cán bộ. Cột Mã GV để trống nếu thêm mới. "
        "Điền thời gian theo định dạng: từ ngày…đến ngày…. Cột Chức vụ/Chức danh chọn từ danh sách."
    )
    ws["A2"].font = instruction_font
    ws["A2"].fill = instruction_fill
    ws["A2"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[2].height = 30

    headers = [
        "Mã GV",
        "Họ tên",
        "Đơn vị công tác",
        "Chức vụ",
        "Thời gian bổ nhiệm chức vụ (từ ngày ….)",
        "Chức danh",
        "Thời gian bổ nhiệm chức danh (từ ngày … )",
        "Thời gian đi học (từ ngày….đến ngày….)",
        "Thời gian đi thực tế(từ ngày….đến ngày….)",
        "Thời gian nghỉ có phép (từ ngày….đến ngày….)",
    ]

    ws.row_dimensions[4].height = 30
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    start_row = 5
    for idx, t in enumerate(teachers):
        current_row = start_row + idx
        ws.row_dimensions[current_row].height = 22

        ws.cell(row=current_row, column=1, value=t["teacher_id"])
        ws.cell(row=current_row, column=2, value=t["name"])
        dept_cell = ws.cell(row=current_row, column=3, value=dept_name)
        dept_cell.protection = Protection(locked=True)
        ws.cell(row=current_row, column=4, value=t["role"] or "")
        ws.cell(row=current_row, column=5, value=str(t["role_date"] or ""))
        ws.cell(row=current_row, column=6, value=t["title"] or "")
        ws.cell(row=current_row, column=7, value=str(t["title_date"] or ""))
        ws.cell(row=current_row, column=8, value="")
        ws.cell(row=current_row, column=9, value="")
        ws.cell(row=current_row, column=10, value="")

        for col_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx != 3:
                cell.protection = Protection(locked=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    extra_rows = 50
    start_extra = start_row + len(teachers)
    for idx in range(extra_rows):
        current_row = start_extra + idx
        ws.row_dimensions[current_row].height = 22

        dept_cell = ws.cell(row=current_row, column=3, value=dept_name)
        dept_cell.protection = Protection(locked=True)

        for col_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx != 3:
                cell.protection = Protection(locked=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    max_row = start_extra + extra_rows - 1

    id_dv = DataValidation(type="whole", allow_blank=True)
    id_dv.prompt = "Mã GV là số nguyên duy nhất. Để trống nếu thêm mới cán bộ."
    id_dv.promptTitle = "Mã Giáo Viên"
    ws.add_data_validation(id_dv)
    id_dv.add(f"A5:A{max_row}")

    role_dv = DataValidation(
        type="list", formula1=f"Metadata!$C$1:$C${len(roles_list_db)}", allow_blank=True
    )
    role_dv.prompt = "Chọn chức vụ lãnh đạo (nếu có)"
    role_dv.promptTitle = "Chức vụ"
    ws.add_data_validation(role_dv)
    role_dv.add(f"D5:D{max_row}")

    title_dv = DataValidation(
        type="list", formula1=f"Metadata!$A$1:$A${len(titles_db)}", allow_blank=True
    )
    title_dv.prompt = "Chọn chức danh giảng dạy phù hợp"
    title_dv.promptTitle = "Chức danh"
    ws.add_data_validation(title_dv)
    title_dv.add(f"F5:F{max_row}")

    col_widths = {1: 10, 2: 22, 3: 22, 4: 18, 5: 26, 6: 18, 7: 26, 8: 26, 9: 26, 10: 26}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_activities_template(dept_name, timeframe_name):
    conn = get_connection()
    cursor = conn.cursor()
    # Fetch teachers in department (just their IDs for validation)
    cursor.execute(
        """
        SELECT t.id FROM teachers t
        WHERE (SELECT value_text FROM teacher_role_history WHERE teacher_id = t.id AND record_type = 'DEPARTMENT' ORDER BY start_date DESC LIMIT 1) = ?
        ORDER BY t.name
    """,
        (dept_name,),
    )
    teacher_ids = [str(r["id"]) for r in cursor.fetchall()]

    # Fetch activity types per category
    cursor.execute(
        "SELECT name FROM activity_types WHERE category = 'Giảng dạy' ORDER BY name"
    )
    act_hdcm = [r["name"] for r in cursor.fetchall()]

    cursor.execute(
        "SELECT name FROM activity_types WHERE category = 'NCKH' ORDER BY name"
    )
    act_nckh = [r["name"] for r in cursor.fetchall()]

    cursor.execute(
        "SELECT name FROM activity_types WHERE category = 'Nhiệm vụ khác' ORDER BY name"
    )
    act_nvk = [r["name"] for r in cursor.fetchall()]
    conn.close()

    if not teacher_ids:
        teacher_ids = ["Chưa cài đặt"]
    if not act_hdcm:
        act_hdcm = ["Chưa cài đặt"]
    if not act_nckh:
        act_nckh = ["Chưa cài đặt"]
    if not act_nvk:
        act_nvk = ["Chưa cài đặt"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Populate Metadata sheet
    meta_ws = wb.create_sheet(title="Metadata")
    meta_ws.sheet_state = "hidden"

    for idx, tid in enumerate(teacher_ids, 1):
        meta_ws.cell(row=idx, column=1, value=tid)
    for idx, act in enumerate(act_hdcm, 1):
        meta_ws.cell(row=idx, column=2, value=act)
    for idx, act in enumerate(act_nckh, 1):
        meta_ws.cell(row=idx, column=3, value=act)
    for idx, act in enumerate(act_nvk, 1):
        meta_ws.cell(row=idx, column=4, value=act)

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="1F5F3F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    instruction_font = Font(name=font_family, size=10, italic=True, color="555555")

    header_fill = PatternFill(
        start_color="1F5F3F", end_color="1F5F3F", fill_type="solid"
    )
    instruction_fill = PatternFill(
        start_color="EAF2EC", end_color="EAF2EC", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def create_sheet(
        title, headers, instruction, act_meta_col, act_meta_len, bool_cols, widths_dict
    ):
        ws = wb.create_sheet(title=title)

        last_col_letter = get_column_letter(len(headers))
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = f"MẪU NHẬP {title.upper()} - NĂM HỌC: {timeframe_name.upper()}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = instruction
        ws["A2"].font = instruction_font
        ws["A2"].fill = instruction_fill
        ws["A2"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[2].height = 30

        ws.row_dimensions[4].height = 30
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = h_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

        num_rows = 100
        start_row = 5
        for idx in range(num_rows):
            current_row = start_row + idx
            ws.row_dimensions[current_row].height = 22

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.protection = Protection(locked=False)
                cell.alignment = (
                    Alignment(horizontal="center", vertical="center")
                    if col_idx in (1, 2, 3, 4) or col_idx in bool_cols
                    else Alignment(vertical="center")
                )

        max_row = start_row + num_rows - 1

        if teacher_ids:
            t_formula = f"Metadata!$A$1:$A${len(teacher_ids)}"
            t_col = get_column_letter(1)
            t_range = f"{t_col}5:{t_col}{max_row}"
            t_dv = DataValidation(type="list", formula1=t_formula, allow_blank=True)
            t_dv.prompt = "Chọn hoặc nhập Mã Giáo viên hợp lệ"
            t_dv.promptTitle = "Mã Giáo viên"
            ws.add_data_validation(t_dv)
            t_dv.add(t_range)

        if act_meta_len > 0:
            act_formula = f"Metadata!${act_meta_col}$1:${act_meta_col}${act_meta_len}"
            act_col = get_column_letter(2)
            act_range = f"{act_col}5:{act_col}{max_row}"
            act_dv = DataValidation(type="list", formula1=act_formula, allow_blank=True)
            act_dv.prompt = "Chọn tên loại hoạt động từ danh sách"
            act_dv.promptTitle = "Loại hoạt động"
            ws.add_data_validation(act_dv)
            act_dv.add(act_range)

        if bool_cols:
            for b_col in bool_cols:
                c_letter = get_column_letter(b_col)
                c_range = f"{c_letter}5:{c_letter}{max_row}"
                yn_dv = DataValidation(
                    type="list", formula1='"Có,Không"', allow_blank=True
                )
                yn_dv.prompt = "Chọn Có hoặc Không"
                yn_dv.promptTitle = "Lựa chọn"
                ws.add_data_validation(yn_dv)
                yn_dv.add(c_range)

        # Date prompt
        date_dv = DataValidation(type="date", allow_blank=True)
        date_dv.prompt = "Nhập ngày thực hiện dạng YYYY-MM-DD"
        date_dv.promptTitle = "Ngày thực hiện"
        ws.add_data_validation(date_dv)
        date_dv.add(f"C5:C{max_row}")

        # Quantity prompt
        qty_dv = DataValidation(type="decimal", allow_blank=True)
        qty_dv.prompt = "Nhập số lượng thực hiện (số dương lớn hơn 0)"
        qty_dv.promptTitle = "Số lượng"
        ws.add_data_validation(qty_dv)
        qty_dv.add(f"D5:D{max_row}")

        for c, w in widths_dict.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.freeze_panes = "A5"

    create_sheet(
        title="Hoạt động chuyên môn",
        headers=[
            "Mã GV",
            "Tên loại hoạt động",
            "Ngày thực hiện",
            "Số lượng",
            "Cấp lớp",
            "Loại lớp",
            "Số học viên",
            "Giảng dạy tiếng nước ngoài",
            "Ghi chú",
        ],
        instruction="Nhập mã GV. Cột Giảng dạy tiếng nước ngoài chọn 'Có' hoặc 'Không'.",
        act_meta_col="B",
        act_meta_len=len(act_hdcm),
        bool_cols=[8],
        widths_dict={1: 15, 2: 30, 3: 16, 4: 10, 5: 16, 6: 16, 7: 12, 8: 22, 9: 22},
    )

    create_sheet(
        title="NCKH",
        headers=[
            "Mã GV",
            "Tên loại hoạt động",
            "Ngày thực hiện",
            "Số lượng",
            "Cấp đề tài",
            "Tác giả chính",
            "Ghi chú",
        ],
        instruction="Nhập mã GV. Cột Tác giả chính chọn 'Có' hoặc 'Không'.",
        act_meta_col="C",
        act_meta_len=len(act_nckh),
        bool_cols=[6],
        widths_dict={1: 15, 2: 30, 3: 16, 4: 10, 5: 16, 6: 16, 7: 22},
    )

    create_sheet(
        title="Nhiệm vụ khác",
        headers=[
            "Mã GV",
            "Tên loại hoạt động",
            "Ngày thực hiện",
            "Số lượng",
            "Ghi chú",
        ],
        instruction="Nhập mã GV và các nhiệm vụ khác.",
        act_meta_col="D",
        act_meta_len=len(act_nvk),
        bool_cols=[],
        widths_dict={1: 15, 2: 30, 3: 16, 4: 10, 5: 30},
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
