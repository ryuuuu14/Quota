import os
import sys

if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

os.environ['DB_PATH'] = 'test_bulk.sqlite'

import unittest
import pandas as pd
import openpyxl
from io import BytesIO

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), 'src')))

from database import init_db, get_connection
from pipeline.templates import generate_excel_template, ALLOWED_LOAI
from pipeline.calculator import calculate_rows, aggregate_by_teacher, lookup_he_so_loai


def validate_excel_data_helper(file_bytes):
    from pipeline.importer import parse_excel_to_df
    from pipeline.validator import validate_schedule_data
    import pandas as pd
    try:
        df = parse_excel_to_df(file_bytes, header_row=3)
    except Exception as e:
        return False, [str(e)], None
    
    # check headers
    expected_headers = [
        "Mã GV (Khóa)", "Họ tên (Khóa)", "Chức danh (Khóa)", "Đơn vị (Khóa)",
        "Tên môn học", "Loại", "Nhóm", "Sỉ số", "Tiết quy đổi",
        "Hệ số tín chỉ", "Ghi chú"
    ]
    actual_headers = [str(h).strip() if not pd.isna(h) else "" for h in df.columns]
    if len(actual_headers) < 11:
        return False, [f"Cần 11 cột, phát hiện {len(actual_headers)} cột."], None
    for i in range(11):
        if actual_headers[i] != expected_headers[i]:
            return False, [f"Cột thứ {i+1} không khớp."], None
            
    conn = get_connection()
    raw_errors = validate_schedule_data(df, conn)
    conn.close()
    
    if raw_errors:
        return False, [f"Dòng {r}: {m}" for _, r, m in raw_errors], None
        
    # Clean and rename
    df.columns = [
        "teacher_id", "teacher_name", "chuc_danh", "don_vi",
        "subject_name", "loai", "nhom", "si_so", "tiet_quy_doi",
        "he_so_tin_chi", "ghi_chu"
    ]
    from pipeline.validator import safe_float
    df["teacher_id"] = df["teacher_id"].apply(lambda x: int(float(x)))
    df["loai"] = df["loai"].astype(str).str.strip().str.upper()
    df["si_so"] = df["si_so"].apply(lambda x: int(safe_float(x)))
    df["tiet_quy_doi"] = df["tiet_quy_doi"].apply(safe_float)
    df["he_so_tin_chi"] = df["he_so_tin_chi"].apply(safe_float)
    return True, [], df


class TestBulkImport(unittest.TestCase):
    def setUp(self):
        if os.path.exists('test_bulk.sqlite'):
            try:
                os.remove('test_bulk.sqlite')
            except:
                pass
        init_db()
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR REPLACE INTO timeframes (id, name, start_date, end_date) "
            "VALUES (1, 'Năm học 2025-2026', '2025-09-01', '2026-06-30')"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teachers (id, name, subject_group, employment_type) "
            "VALUES (1, 'Nguyễn Văn A', 'Chính trị/Nghiệp vụ', 'TEACHER')"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teachers (id, name, subject_group, employment_type) "
            "VALUES (2, 'Trần Thị B', 'Tự nhiên/Kỹ thuật', 'TEACHER')"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teacher_role_history "
            "(teacher_id, record_type, value_text, start_date, end_date) "
            "VALUES (1, 'TITLE', 'Giảng viên', '2025-09-01', NULL)"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teacher_role_history "
            "(teacher_id, record_type, value_text, start_date, end_date) "
            "VALUES (1, 'DEPARTMENT', 'Khoa Xã hội', '2025-09-01', NULL)"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teacher_role_history "
            "(teacher_id, record_type, value_text, start_date, end_date) "
            "VALUES (2, 'TITLE', 'Giảng viên chính', '2025-09-01', NULL)"
        )
        cursor.execute(
            "INSERT OR REPLACE INTO teacher_role_history "
            "(teacher_id, record_type, value_text, start_date, end_date) "
            "VALUES (2, 'DEPARTMENT', 'Khoa Kỹ thuật', '2025-09-01', NULL)"
        )
        conn.commit()
        conn.close()

    def tearDown(self):
        if os.path.exists('test_bulk.sqlite'):
            try:
                os.remove('test_bulk.sqlite')
            except:
                pass

    def test_template_generation(self):
        template_bytes = generate_excel_template("Năm học 2025-2026")
        self.assertIsNotNone(template_bytes)
        self.assertTrue(len(template_bytes) > 0)

        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        sheet = wb.active
        self.assertIn("Nhập", wb.sheetnames[0])

        headers = [cell.value for cell in sheet[4]]
        self.assertIn("Mã GV (Khóa)", headers)
        self.assertIn("Họ tên (Khóa)", headers)
        self.assertIn("Chức danh (Khóa)", headers)
        self.assertIn("Đơn vị (Khóa)", headers)
        self.assertIn("Tên môn học", headers)
        self.assertIn("Loại", headers)
        self.assertIn("Sỉ số", headers)
        self.assertIn("Tiết quy đổi", headers)
        self.assertIn("Hệ số tín chỉ", headers)
        self.assertEqual(len(headers), 11)

        row5 = [cell.value for cell in sheet[5]]
        self.assertEqual(row5[0], 1)
        self.assertEqual(row5[1], "Nguyễn Văn A")

    def test_lookup_he_so_loai(self):
        self.assertEqual(lookup_he_so_loai("LT", 30), 1.0)
        self.assertEqual(lookup_he_so_loai("LT", 45), 1.2)
        self.assertEqual(lookup_he_so_loai("LT", 70), 1.4)
        self.assertEqual(lookup_he_so_loai("LT", 100), 1.5)
        self.assertEqual(lookup_he_so_loai("TH", 30), 1.0)
        self.assertEqual(lookup_he_so_loai("TH", 50), 1.2)
        self.assertEqual(lookup_he_so_loai("TH", 65), 1.4)
        self.assertEqual(lookup_he_so_loai("TH", 80), 1.5)
        self.assertEqual(lookup_he_so_loai("NN_CNTT", 20), 1.0)
        self.assertEqual(lookup_he_so_loai("NN_CNTT", 30), 1.2)
        self.assertEqual(lookup_he_so_loai("THẠC SĨ", 30), 1.3)
        self.assertEqual(lookup_he_so_loai("THẠC SĨ", 60), 1.5)
        self.assertEqual(lookup_he_so_loai("TIẾN SĨ", 10), 2.0)
        self.assertEqual(lookup_he_so_loai("LLCT TRUNG CẤP", 40), 1.0)
        self.assertEqual(lookup_he_so_loai("LLCT TRUNG CẤP", 55), 1.2)
        self.assertEqual(lookup_he_so_loai("LLCT CAO CẤP", 40), 1.3)
        self.assertEqual(lookup_he_so_loai("LLCT CAO CẤP", 55), 1.5)
        self.assertEqual(lookup_he_so_loai("BỒI DƯỠNG", 100), 1.0)

    def test_full_flow(self):
        template_bytes = generate_excel_template("Năm học 2025-2026")

        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        sheet = wb.active

        sheet["E5"] = "Toán cao cấp"
        sheet["F5"] = "LT"
        sheet["G5"] = "01"
        sheet["H5"] = 45
        sheet["I5"] = 45
        sheet["J5"] = 2.0

        sheet["E6"] = "Cấu trúc dữ liệu"
        sheet["F6"] = "LT"
        sheet["G6"] = "02"
        sheet["H6"] = 60
        sheet["I6"] = 30
        sheet["J6"] = 2.0

        out = BytesIO()
        wb.save(out)
        modified_bytes = out.getvalue()

        is_valid, errors, parsed_df = validate_excel_data_helper(modified_bytes)
        self.assertTrue(is_valid, msg=f"Validation failed: {errors}")
        self.assertEqual(len(errors), 0)
        self.assertEqual(len(parsed_df), 2)

        df_calc = calculate_rows(parsed_df)
        self.assertIn("he_so_lop_dong", df_calc.columns)
        self.assertIn("tiet_thuc_day", df_calc.columns)

        agg = aggregate_by_teacher(df_calc)
        self.assertEqual(len(agg), 2)
        t1 = agg[agg["teacher_id"] == 1]
        self.assertFalse(t1.empty)

    def test_aggregate_totals_import(self):
        from pipeline.validator import validate_aggregate_totals_data
        from pipeline.differ import diff_aggregate_totals

        # Create df with valid columns but some invalid data to test validation
        df_invalid = pd.DataFrame([
            {
                "Mã GV": "999",  # non-existent teacher
                "Tổng GC thực hiện": "abc",
                "NCKH thực hiện": "50.0",
                "Số giờ miễn giảm": "20.0",
                "Định mức GC": "270.0",
                "Ghi chú": ""
            }
        ])

        conn = get_connection()
        errors = validate_aggregate_totals_data(df_invalid, conn)
        self.assertTrue(len(errors) > 0, "Should detect errors (invalid number, non-existent teacher)")

        # Create df with valid data
        df_valid = pd.DataFrame([
            {
                "Mã GV": "1",  # Nguyen Van A (exists)
                "Tổng GC thực hiện": "320.5",
                "NCKH thực hiện": "80.0",
                "Số giờ miễn giảm": "15.0",
                "Định mức GC": "270.0",
                "Ghi chú": "Ghi đè thủ công"
            }
        ])

        errors_valid = validate_aggregate_totals_data(df_valid, conn)
        self.assertEqual(len(errors_valid), 0, f"Valid df should have 0 errors, got: {errors_valid}")

        # Test diff
        df_diff = diff_aggregate_totals(df_valid, conn, "Năm học 2025-2026")
        self.assertEqual(len(df_diff), 1)
        self.assertEqual(df_diff.iloc[0]["diff_marker"], "NEW")

        # Insert record into teacher_calculated_totals to simulate pre-existing override
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO teacher_calculated_totals (teacher_id, timeframe_id, tong_gc_da_thuc_hien, is_override)
            VALUES (1, 1, 150.0, 1)
        """)
        conn.commit()

        # Diff again
        df_diff_update = diff_aggregate_totals(df_valid, conn, "Năm học 2025-2026")
        self.assertEqual(df_diff_update.iloc[0]["diff_marker"], "UPDATE")

        conn.close()

    def test_legacy_xls_support(self):
        from unittest.mock import patch
        from pipeline.importer import get_excel_sheet_names

        # Mock openpyxl to fail, triggering the xlrd fallback
        with patch("openpyxl.load_workbook") as mock_load:
            mock_load.side_effect = Exception("Unsupported format")
            
            with patch("xlrd.open_workbook") as mock_xlrd:
                class MockWorkbook:
                    def sheet_names(self):
                        return ["Sheet1", "metadata", "Sheet2"]
                mock_xlrd.return_value = MockWorkbook()
                
                sheets = get_excel_sheet_names(b"fake_xls_bytes")
                self.assertEqual(sheets, ["Sheet1", "Sheet2"])
                mock_xlrd.assert_called_once_with(file_contents=b"fake_xls_bytes")

    def test_xls_parser_integration(self):
        from unittest.mock import patch, ANY
        from pipeline.importer import get_excel_headers, parse_excel_to_df

        # When parsing legacy .xls, pandas automatically routes via the xlrd engine.
        # We mock pd.read_excel to simulate parsing the excel file.
        with patch("pandas.read_excel") as mock_read:
            mock_read.return_value = pd.DataFrame(columns=["Mã GV", "Tổng GC thực hiện"])
            
            headers = get_excel_headers(b"xls_bytes", sheet_name="Sheet1", header_row=0)
            self.assertEqual(headers, ["Mã GV", "Tổng GC thực hiện"])
            mock_read.assert_called_once_with(ANY, sheet_name="Sheet1", header=0, nrows=1)

            mock_read.reset_mock()
            mock_read.return_value = pd.DataFrame([{"Mã GV": "1", "Tổng GC thực hiện": "150.0"}])
            df = parse_excel_to_df(b"xls_bytes", header_row=0, sheet_name="Sheet1")
            self.assertEqual(len(df), 1)
            self.assertEqual(df.iloc[0]["Mã GV"], "1")

    def test_mapping_templates_db_crud(self):
        from pipeline.mapping_templates import load_mapping_templates, save_mapping_template, delete_mapping_template
        
        # Verify initial load
        tpls = load_mapping_templates()
        self.assertIsInstance(tpls, dict)

        # Save a new template
        mapping = {"Mã GV": "Mã giảng viên", "Số lượng": "Giờ"}
        save_mapping_template("TestTemplate", mapping)

        # Load and verify
        tpls = load_mapping_templates()
        self.assertIn("TestTemplate", tpls)
        self.assertEqual(tpls["TestTemplate"], mapping)

        # Delete and verify
        delete_mapping_template("TestTemplate")
        tpls = load_mapping_templates()
        self.assertNotIn("TestTemplate", tpls)


if __name__ == "__main__":
    unittest.main()
